import os 
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except ImportError:
    pass
from functools import wraps
from flask import Flask, jsonify, request, abort, render_template, redirect, url_for, flash, send_file, session, send_from_directory, make_response, Response
from queue import Queue
import io
import traceback
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, or_
from sqlalchemy.exc import IntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
# pandas removido para aligerar el servidor
from fpdf import FPDF
from datetime import datetime, date, timedelta, time
import json
import socket
import time

intentos_login = {}
import csv
from urllib.parse import urlparse

import pytz
# --- Configuración de Zona Horaria (Argentina UTC-3) ---
def hora_argentina():
    tz = pytz.timezone('America/Argentina/Buenos_Aires')
    return datetime.now(tz).replace(tzinfo=None)

ultima_actualizacion_precios = hora_argentina()
ultima_actualizacion_catalogo = hora_argentina().isoformat()

def actualizar_version_catalogo():
    global ultima_actualizacion_catalogo
    ultima_actualizacion_catalogo = hora_argentina().isoformat()

# ─── Configuración ────────────────────────────────────────────
DB_PATH = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'tienda.db')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())
app.config['SECRET_KEY'] = app.secret_key
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)
app.config['TEMPLATES_AUTO_RELOAD'] = True

def es_accesible_bd_nube(uri_nube):
    if not uri_nube: return False
    
    # 1. Comprobación ultra-rápida de red física (sin DNS) a una IP fija (Cloudflare DNS)
    # Esto falla inmediatamente en microsegundos si el cable de red o Wi-Fi está apagado.
    try:
        socket.setdefaulttimeout(1.0)
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect(("1.1.1.1", 53))
        s.close()
    except Exception:
        print("[DETECTADO] Red física desconectada. Evitando bloqueos de DNS.")
        return False

    # 2. Resolución de DNS en un hilo secundario con un timeout de 1.5 segundos
    # para evitar que el gethostbyname síncrono de Windows congele el servidor Flask.
    import threading
    url = urlparse(uri_nube)
    hostname = url.hostname
    port = url.port or 5432
    if not hostname: return False
    
    dns_res = [None]
    def lookup():
        try:
            dns_res[0] = socket.gethostbyname(hostname)
        except Exception:
            pass
            
    t = threading.Thread(target=lookup)
    t.daemon = True
    t.start()
    t.join(timeout=1.5)
    
    if not dns_res[0]:
        print(f"[CONEXION] No se pudo resolver el host de la nube: {hostname}")
        return False
        
    # 3. Comprobación rápida de conexión TCP al puerto
    try:
        socket.create_connection((dns_res[0], port), timeout=1.5)
        return True
    except Exception as e:
        print(f"[CONEXION] Falló conexión TCP al host {dns_res[0]}:{port}: {e}")
        return False

# ─── Configuración de Base de Datos con Detección de Entorno Automática ───
uri_nube = os.environ.get('DATABASE_URL')
if not uri_nube:
    raise RuntimeError("DATABASE_URL no configurada en las variables de entorno.")
if uri_nube.startswith('postgres://'):
    uri_nube = uri_nube.replace('postgres://', 'postgresql://', 1)

# Detectar si estamos en el entorno de Render (producción)
is_render_production = os.environ.get('RENDER') == 'true'

try:
    if is_render_production:
        if not uri_nube:
            raise Exception("DATABASE_URL no definida en entorno de Render.")
        app.config['SQLALCHEMY_DATABASE_URI'] = uri_nube
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
            "pool_pre_ping": True,
            "pool_recycle": 300,
            "connect_args": {
                "connect_timeout": 3  # Timeout estricto de conexión de 3 segundos para Postgres
            }
        }
        print("[BACKEND] -> Conectado a PostgreSQL en Render (Nube)")
    else:
        # Modo híbrido local: intentar conectar a base de datos de Render si está disponible de forma rápida
        if not uri_nube or urlparse(uri_nube).hostname in ['localhost', '127.0.0.1'] or not es_accesible_bd_nube(uri_nube):
            raise Exception("DATABASE_URL remota no accesible o no configurada para el entorno local.")
        
        app.config['SQLALCHEMY_DATABASE_URI'] = uri_nube
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
            "pool_pre_ping": True,
            "pool_recycle": 300,
            "connect_args": {
                "connect_timeout": 3  # Timeout estricto de conexión de 3 segundos para Postgres
            }
        }
        print("[BACKEND] -> Conectado a PostgreSQL en Render (Nube)")
except Exception as e:
    # Asegurar fallback absoluto a SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        "pool_pre_ping": True,
        "pool_recycle": 300,
        "connect_args": {
            "timeout": 15  # SQLite busy_timeout de 15 segundos
        }
    }
    print(f"[BACKEND] -> ¡SIN INTERNET! Operando local con tienda.db (Detalle: {e})")

CORS(app)
db = SQLAlchemy(app)

def es_offline():
    return 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', '')

def login_requerido(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_autenticado'):
            if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"ok": False, "error": "No autorizado"}), 401
            return redirect('/')
        return f(*args, **kwargs)
    return decorated_function


# ─── Modelos SQLAlchemy ───────────────────────────────────────
class Usuario(db.Model):
    __tablename__ = 'Usuarios'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)

    def __init__(self, **kwargs):
        super(Usuario, self).__init__(**kwargs)

class Categoria(db.Model):
    __tablename__ = 'categoria'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False)

    def __init__(self, **kwargs):
        super(Categoria, self).__init__(**kwargs)

class Producto(db.Model):
    __tablename__ = 'productos'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    descripcion = db.Column(db.Text)
    precio_lista_1 = db.Column(db.Float, nullable=False) # Lista 1
    precio_lista_2 = db.Column(db.Float, nullable=True, default=0.0) # Lista 2
    precio_lista_3 = db.Column(db.Float, nullable=True, default=0.0) # Lista 3
    precio_anterior = db.Column(db.Float, nullable=True)
    imagen = db.Column(db.String(255), default='')
    imagen_url = db.Column(db.Text, default='')
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria.id'))
    categoria_rel = db.relationship('Categoria', backref='productos')
    stock = db.Column(db.Integer, default=0)
    activo = db.Column(db.Integer, default=1)
    
    favorito = db.Column(db.Boolean, default=False)
    permitir_sin_stock = db.Column(db.Boolean, default=True)
    ventas_totales = db.Column(db.Integer, default=0)
    codigo_barra = db.Column(db.String(100), nullable=True)

    # ─── Descuento por Volumen / Mayorista ───
    descuento_volumen_activo = db.Column(db.Boolean, default=False)
    cantidad_minima_descuento = db.Column(db.Integer, nullable=True)
    porcentaje_descuento_volumen = db.Column(db.Float, nullable=True)

    sincronizado = db.Column(db.Boolean, default=True, nullable=False)
    ultima_actualizacion = db.Column(db.DateTime, default=hora_argentina, onupdate=hora_argentina)

    def __init__(self, **kwargs):
        super(Producto, self).__init__(**kwargs)

    @property
    def precio(self):
        return self.precio_lista_1

    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'precio': self.precio_lista_1,
            'precio_lista_1': self.precio_lista_1,
            'precio_lista_2': self.precio_lista_2 or self.precio_lista_1,
            'precio_lista_3': self.precio_lista_3 or self.precio_lista_1,
            'precio_anterior': self.precio_anterior,
            'imagen': self.imagen_url or self.imagen,
            'imagen_url': self.imagen_url,
            'categoria': self.categoria_rel.nombre if self.categoria_rel else 'General',
            'stock': self.stock,
            'activo': self.activo,
            'favorito': self.favorito,
            'permitir_sin_stock': self.permitir_sin_stock,
            'ventas_totales': self.ventas_totales,
            'codigo_barra': self.codigo_barra or '',
            'descuento_volumen_activo': self.descuento_volumen_activo,
            'cantidad_minima_descuento': self.cantidad_minima_descuento,
            'porcentaje_descuento_volumen': self.porcentaje_descuento_volumen,
            'sincronizado': self.sincronizado,
            'ultima_actualizacion': self.ultima_actualizacion.isoformat() if self.ultima_actualizacion else None
        }

# ─── Modelos de Clientes y Ventas ───────────────────────────────────────
class Cliente(db.Model):
    __tablename__ = 'clientes'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    cuit = db.Column(db.String(20), nullable=True, default='', unique=True)
    condicion_iva = db.Column(db.String(50), nullable=True, default='Consumidor Final')
    telefono = db.Column(db.String(30), nullable=True, default='')
    direccion = db.Column(db.String(255), nullable=True, default='')
    descuento = db.Column(db.Float, default=0.0)
    descuento_fijo = db.Column(db.Float, default=0.0)
    saldo = db.Column(db.Float, default=0.0)
    limite_credito = db.Column(db.Float, default=0.0)
    activo = db.Column(db.Integer, default=1)
    ventas = db.relationship('Venta', backref='cliente', lazy=True)

    def __init__(self, **kwargs):
        super(Cliente, self).__init__(**kwargs)

    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'cuit': self.cuit or '',
            'condicion_iva': self.condicion_iva or 'CF',
            'telefono': self.telefono or '',
            'direccion': self.direccion or '',
            'descuento_fijo': self.descuento_fijo or 0.0,
            'descuento': self.descuento or 0.0,
            'saldo': self.saldo or 0.0,
            'limite_credito': self.limite_credito or 0.0
        }

class Venta(db.Model):
    __tablename__ = 'ventas'
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('clientes.id'), nullable=True)
    fecha = db.Column(db.DateTime, default=hora_argentina)
    total = db.Column(db.Float, default=0.0)
    detalle_json = db.Column(db.Text, default='[]')
    lista_precios = db.Column(db.Integer, default=1)
    tipo = db.Column(db.String(50), default='local')
    metodo_pago = db.Column(db.String(100), nullable=True)
    pago_efectivo = db.Column(db.Float, default=0.0)
    pago_transferencia = db.Column(db.Float, default=0.0)
    pago_debito = db.Column(db.Float, default=0.0)
    pago_cc = db.Column(db.Float, default=0.0)
    entregado = db.Column(db.Float, default=0.0)
    sincronizado = db.Column(db.Boolean, default=True, nullable=False)
    ultima_actualizacion = db.Column(db.DateTime, default=hora_argentina, onupdate=hora_argentina)
    detalles = db.relationship('DetalleVenta', backref='venta', lazy=True, cascade="all, delete-orphan")

    def __init__(self, **kwargs):
        super(Venta, self).__init__(**kwargs)

class Gasto(db.Model):
    __tablename__ = 'gastos'
    id = db.Column(db.Integer, primary_key=True)
    descripcion = db.Column(db.String(255), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.DateTime, default=hora_argentina)
    categoria = db.Column(db.String(100), default='General')
    tipo = db.Column(db.String(20), default='Egreso') # 'Egreso' o 'Ingreso'

    def __init__(self, **kwargs):
        super(Gasto, self).__init__(**kwargs)

class DetalleVenta(db.Model):
    __tablename__ = 'detalle_ventas'
    id = db.Column(db.Integer, primary_key=True)
    venta_id = db.Column(db.Integer, db.ForeignKey('ventas.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=True)
    nombre_producto = db.Column(db.String(150), nullable=False)
    cantidad = db.Column(db.Integer, default=1)
    precio_unitario = db.Column(db.Float, default=0.0)
    descuento_porcentaje = db.Column(db.Float, default=0.0)
    subtotal = db.Column(db.Float, default=0.0)

    def __init__(self, **kwargs):
        super(DetalleVenta, self).__init__(**kwargs)

    def to_dict(self):
        f = self.fecha
        if isinstance(f, str):
            try:
                from datetime import datetime
                f = datetime.fromisoformat(f.replace('Z', '+00:00'))
            except:
                pass
        
        return {
            'id': self.id,
            'descripcion': self.descripcion,
            'monto': self.monto,
            'fecha': f.strftime('%H:%M') if hasattr(f, 'strftime') else '--:--',
            'categoria': self.categoria,
            'tipo': self.tipo
        }

class CajaDiaria(db.Model):
    __tablename__ = 'cajas_diarias'
    id = db.Column(db.Integer, primary_key=True)
    fecha_apertura = db.Column(db.DateTime, default=hora_argentina)
    monto_inicial = db.Column(db.Float, nullable=False)
    monto_final = db.Column(db.Float, nullable=True)
    fecha_cierre = db.Column(db.DateTime, nullable=True)
    estado = db.Column(db.String(20), default='Abierta') # 'Abierta' o 'Cerrada'
    turno = db.Column(db.String(20), default='Mañana') # 'Mañana' o 'Tarde'

    def __init__(self, **kwargs):
        super(CajaDiaria, self).__init__(**kwargs)
    
    def to_dict(self):
        return {
            'id': self.id,
            'monto_inicial': self.monto_inicial,
            'monto_final': self.monto_final,
            'estado': self.estado,
            'turno': self.turno,
            'fecha_apertura': self.fecha_apertura.strftime('%Y-%m-%d %H:%M') if self.fecha_apertura else '',
            'fecha_cierre': self.fecha_cierre.strftime('%Y-%m-%d %H:%M') if self.fecha_cierre else ''
        }

# ─── Variables para Server-Sent Events (SSE) ─────────────────────────────────
# Lista global para mantener las colas de los clientes de SSE (cajas)
sse_clients = []

@app.route('/api/stream-actualizaciones')
def stream_actualizaciones():
    def event_stream():
        q = Queue()
        sse_clients.append(q)
        try:
            while True:
                data = q.get()
                yield f"data: {json.dumps(data)}\n\n"
        except GeneratorExit:
            if q in sse_clients:
                sse_clients.remove(q)
                
    return Response(event_stream(), mimetype="text/event-stream")

@app.route('/api/enviar-aviso', methods=['POST'])
@login_requerido
def enviar_aviso():
    if not session.get('admin_autenticado'):
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401
        
    data = request.json or {}
    modificados = data.get('modificados', [])
    
    if not modificados:
        return jsonify({'ok': False, 'mensaje': 'No hay artículos para notificar'}), 400
        
    # Enviar a todos los clientes SSE conectados
    for q in sse_clients:
        q.put({"tipo": "actualizacion_precios", "modificados": modificados})
        
    return jsonify({'ok': True, 'mensaje': 'Aviso enviado a las cajas correctamente.'})

# ─── INICIALIZACIÓN CRÍTICA (Render/Gunicorn compatible) ──────
with app.app_context():
    db.create_all()
    try:
        db.session.execute(text("ALTER TABLE cajas_diarias ADD COLUMN turno VARCHAR(20) DEFAULT 'Mañana';"))
        db.session.commit()
    except:
        db.session.rollback()
    print("Base de datos y tablas inicializadas correctamente.")


def verificar_corte_automatico():
    if request.path.startswith('/static/'):
        return
    try:
        caja = CajaDiaria.query.filter_by(estado='Abierta').order_by(CajaDiaria.id.desc()).first()
        if caja:
            tz = pytz.timezone('America/Argentina/Buenos_Aires')
            ahora = datetime.now(tz)
            apertura = caja.fecha_apertura
            
            cierre_necesario = False
            if apertura and apertura.date() < ahora.date():
                cierre_necesario = True
            elif caja.turno == 'Mañana' and ahora.hour >= 13 and (apertura and apertura.hour < 13):
                cierre_necesario = True
            elif caja.turno == 'Tarde' and ahora.hour >= 22 and (apertura and apertura.hour < 22):
                cierre_necesario = True
                
            if cierre_necesario:
                caja.estado = 'Cerrada'
                caja.fecha_cierre = hora_argentina()
                db.session.commit()
                print(f"Corte automático ejecutado para caja turno {caja.turno} ID: {caja.id}")
    except Exception as e:
        print(f"Error en corte automático: {e}")

@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ─── API REST (Para el Frontend) ─────────────────────────────
@app.route('/api/estado_conexion', methods=['GET'])
@login_requerido
def estado_conexion():
    return jsonify({"online": not es_offline()})


# ─── RUTA TEMPORAL DE MIGRACIÓN ───
@app.route('/forzar-migracion-db')
@login_requerido
def forzar_migracion_db():
    if not session.get('admin_autenticado'): return redirect('/')
    try:
        from sqlalchemy import text
        db.session.execute(text('ALTER TABLE productos ADD COLUMN IF NOT EXISTS sincronizado BOOLEAN DEFAULT TRUE;'))
        db.session.execute(text('ALTER TABLE productos ADD COLUMN IF NOT EXISTS ultima_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP;'))
        db.session.commit()
        return "<h1>¡Sincronización forzada con SQLAlchemy Exitosa!</h1>", 200
    except Exception as e:
        db.session.rollback()
        return f"<h1>Error al inyectar: {str(e)}</h1><pre>{traceback.format_exc()}</pre>", 500


@app.route('/importar-csv')
@login_requerido
def importar_csv():
    if not session.get('admin_autenticado'): return redirect('/')
    try:
        csv_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'lista_productos.csv')
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                nombre = row.get('Nombre', '').strip()
                if not nombre:
                    continue
                
                try:
                    p1_str = row.get('Lista 1', '').replace('$', '').replace(',', '.').strip()
                    precio_lista_1 = float(p1_str) if p1_str else 0.0
                except ValueError:
                    precio_lista_1 = 0.0
                    
                codigo_barra = row.get('Cod. Barra', '').strip()
                
                producto = Producto(
                    nombre=nombre,
                    precio_lista_1=precio_lista_1,
                    codigo_barra=codigo_barra,
                    activo=1,
                    stock=0,
                    sincronizado=True
                )
                db.session.add(producto)
        
        db.session.commit()
        return "<h1>¡Importación Exitosa! Productos cargados en Supabase.</h1>", 200
    except Exception as e:
        db.session.rollback()
        return f"<h1>Error al importar: {str(e)}</h1>", 500


@app.route('/optimizar-db')
@login_requerido
def optimizar_db():
    if not session.get('admin_autenticado'): return redirect('/')
    try:
        db.session.execute(text('CREATE INDEX IF NOT EXISTS idx_producto_nombre ON productos (nombre);'))
        db.session.execute(text('CREATE INDEX IF NOT EXISTS idx_producto_codigo_barra ON productos (codigo_barra);'))
        db.session.execute(text('CREATE INDEX IF NOT EXISTS idx_producto_activo ON productos (activo);'))
        db.session.commit()
        return "<h1>¡Turbo Activado! Índices creados en Supabase.</h1>", 200
    except Exception as e:
        db.session.rollback()
        return f"<h1>Error al crear índices: {str(e)}</h1><pre>{traceback.format_exc()}</pre>", 500


@app.route('/api/productos', methods=['GET'])
@login_requerido
def get_productos():
    if es_offline():
        print("[SERVIDO LOCAL] -> Ejecutando consulta en tienda.db offline (api/productos)")
    categoria_nombre = request.args.get('categoria', '').strip()
    buscar    = request.args.get('buscar', '').strip()
    orden     = request.args.get('orden', 'id')

    query = Producto.query.filter_by(activo=1)

    if categoria_nombre:
        cat = Categoria.query.filter_by(nombre=categoria_nombre).first()
        if cat:
            query = query.filter_by(categoria_id=cat.id)
        else:
            query = query.filter_by(id=0) # Fuerza vacío si no existe

    if buscar:
        query = query.filter(Producto.nombre.ilike(f'%{buscar}%') | Producto.descripcion.ilike(f'%{buscar}%'))

    if orden == 'precio_asc':
        query = query.order_by(Producto.precio_lista_1.asc())
    elif orden == 'ventas':
        query = query.order_by(Producto.ventas_totales.desc())
    else:
        query = query.order_by(Producto.id.desc())

    productos = query.limit(50).all()
    return jsonify({"productos": [p.to_dict() for p in productos]})

@app.route('/api/productos/catalogo_completo', methods=['GET'])
@login_requerido
def catalogo_completo():
    productos = Producto.query.filter_by(activo=1).order_by(Producto.nombre.asc()).all()
    return jsonify({
        "productos": [
            {
                "id": p.id,
                "nombre": p.nombre,
                "precio_lista_1": p.precio_lista_1,
                "precio_lista_2": p.precio_lista_2 or p.precio_lista_1,
                "precio_lista_3": p.precio_lista_3 or p.precio_lista_1,
                "stock": p.stock,
                "codigo_barra": p.codigo_barra,
                "categoria": p.categoria_rel.nombre if p.categoria_rel else "General"
            } for p in productos
        ]
    })

@app.route('/buscar_productos')
@login_requerido
def buscar_productos():
    if es_offline():
        print("[SERVIDO LOCAL] -> Ejecutando consulta en tienda.db offline (buscar_productos)")
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"productos": []})
    
    # Búsqueda insensible a mayúsculas con ILIKE
    prods = Producto.query.filter(
        Producto.nombre.ilike(f'%{q}%'),
        Producto.activo == 1
    ).limit(30).all()
    
    return jsonify({
        "productos": [p.to_dict() for p in prods]
    })

@app.route('/buscar_por_codigo/<codigo>')
@login_requerido
def buscar_por_codigo(codigo):
    if es_offline():
        print(f"[SERVIDO LOCAL] -> Buscando código '{codigo}' en tienda.db offline (buscar_por_codigo)")
    codigo = codigo.strip()
    
    # Búsqueda multi-variante (coma-separated barcodes)
    try:
        # Filtro rápido en BD para traer candidatos potenciales
        candidatos = Producto.query.filter(
            Producto.codigo_barra.ilike(f'%{codigo}%'),
            Producto.activo == 1
        ).all()
        
        producto_encontrado = None
        # Validación estricta en Python
        for p in candidatos:
            if not p.codigo_barra: continue
            lista_codigos = [c.strip() for c in p.codigo_barra.split(',') if c.strip()]
            if codigo in lista_codigos:
                producto_encontrado = p
                break
        
        p = producto_encontrado
    except Exception:
        p = None
        
    # Fallback: buscar por ID if el código es numérico
    if not p and codigo.isdigit():
        p = Producto.query.filter_by(id=int(codigo), activo=1).first()
        
    if p:
        return jsonify({"ok": True, "producto": p.to_dict()})
    else:
        return jsonify({"ok": False, "mensaje": "Producto no encontrado"}), 404

@app.route('/api/productos/<int:producto_id>', methods=['GET'])
@login_requerido
def get_producto(producto_id):
    if es_offline():
        print(f"[SERVIDO LOCAL] -> Consultando producto ID {producto_id} en tienda.db offline (get_producto)")
    producto = Producto.query.filter_by(id=producto_id, activo=1).first()
    if not producto:
        abort(404, description="Producto no encontrado")
    return jsonify({"ok": True, "producto": producto.to_dict()})

@app.route('/api/categorias', methods=['GET'])
@login_requerido
def get_categorias():
    categorias = db.session.query(Categoria.nombre, db.func.count(Producto.id))\
        .join(Producto, Producto.categoria_id == Categoria.id)\
        .filter(Producto.activo == 1)\
        .group_by(Categoria.nombre).all()
    return jsonify({
        "ok": True,
        "categorias": [{"categoria": c[0], "cantidad": c[1]} for c in categorias]
    })

@app.route('/api/registrar_venta', methods=['POST'])
@login_requerido
def registrar_venta():
    import json as _json
    data = request.json
    if not data:
        return jsonify({"ok": False, "mensaje": "Datos inválidos"}), 400

    # Capturamos la decisión del usuario (por defecto False)
    facturar_afip = data.get('facturar_afip', False) if isinstance(data, dict) else False

    afip_ok = False
    mensaje_afip = ""

    if facturar_afip:
        try:
            # Simulación de llamada a AFIP (Aquí iría afip.py)
            # Si no hay internet, esto lanzará una excepción por timeout
            print("📡 Intentando conectar con servidores de AFIP...")
            # check_afip_status() ...
            afip_ok = True
        except Exception as e:
            print(f"⚠️ Error de conexión AFIP: {e}")
            afip_ok = False
            mensaje_afip = " (Modo Offline: Pendiente de CAE)"

    tipo_comprobante = "Factura C" if afip_ok else "Ticket No Fiscal"
    
    # Acepta dos formatos:
    # A) Lista directa (uso original del carrito público): [{id/producto_id, qty}, ...]
    # B) Objeto con cliente: {cliente_id, items:[{id/producto_id,qty,name,precio_unit}], total}
    if isinstance(data, list):
        items = data
        cliente_id = None
        lista_sel = 1
        general_discount_perc = 0.0
    else:
        items = data.get('items', [])
        cliente_id = data.get('cliente_id')
        try:
            lista_sel = int(data.get('lista_precios', 1))
        except (ValueError, TypeError):
            lista_sel = 1
        try:
            general_discount_perc = float(data.get('general_discount_perc', 0.0))
        except (ValueError, TypeError):
            general_discount_perc = 0.0

    if not items or len(items) == 0:
        return jsonify({"ok": False, "mensaje": "No se puede registrar una venta sin artículos."}), 400

    # ─── Validación Previa de Productos ──────────────────────────────
    for item in items:
        p_id = item.get('producto_id') if item.get('producto_id') is not None else item.get('id')
        if not p_id:
            db.session.rollback()
            return jsonify({"error": "No se especificó un ID de producto válido en los ítems de la venta."}), 400

        try:
            p_id_int = int(p_id)
        except (ValueError, TypeError):
            p_id_int = None

        producto = Producto.query.get(p_id_int) if p_id_int is not None else None
        if not producto:
            db.session.rollback()
            return jsonify({"error": f"El producto ID {p_id} es None (No existe en la BD)"}), 400
        
        valor_activo = getattr(producto, 'activo', 'No tiene atributo activo')
        if valor_activo in [0, False, '0', None]:
            db.session.rollback()
            return jsonify({"error": f"El producto ID {p_id} está inactivo. Valor de activo: {valor_activo}"}), 400

    # Recalcular el total y construir el detalle de forma segura en el backend para evitar forjado de montos (L-04)
    total_recalculado = 0.0
    detalle_recalculado = []

    for item in items:
        p_id = item.get('producto_id') if item.get('producto_id') is not None else item.get('id')
        try:
            qty = int(item.get('qty', 1))
        except (ValueError, TypeError):
            qty = 1
            
        if not p_id or qty <= 0:
            continue
        
        p = Producto.query.get(int(p_id))
        if p:
            # Obtener el precio oficial registrado en la base de datos
            precio_u = p.precio_lista_3 if lista_sel == 3 else (p.precio_lista_2 if lista_sel == 2 else p.precio_lista_1)
            
            # Obtener descuento de línea enviado por el frontend
            try:
                discount_perc = float(item.get('discount_perc', 0.0))
            except (ValueError, TypeError):
                discount_perc = 0.0
            
            # Aplicar descuento de volumen si califica y es mayor que el descuento de línea
            if p.descuento_volumen_activo and p.cantidad_minima_descuento and qty >= p.cantidad_minima_descuento:
                desc_vol = float(p.porcentaje_descuento_volumen or 0.0)
                if desc_vol > discount_perc:
                    discount_perc = desc_vol
            
            precio_final = round(precio_u - (precio_u * (discount_perc / 100.0)), 2)
            subtotal = round(precio_final * qty, 2)
            
            total_recalculado += subtotal
            
            detalle_recalculado.append({
                'producto_id': p.id,
                'nombre': p.nombre,
                'qty': qty,
                'precio_unit': precio_final,
                'discount_perc': discount_perc,
                'subtotal': subtotal
            })

    # Aplicar el descuento general si existe
    if general_discount_perc > 0.0:
        total_recalculado = round(total_recalculado - (total_recalculado * (general_discount_perc / 100.0)), 2)

    total_venta = total_recalculado
    detalle = detalle_recalculado

    from datetime import timedelta
    tiempo_limite = hora_argentina() - timedelta(seconds=60)
    detalle_json_str = _json.dumps(detalle)
    venta_fantasma = Venta.query.filter(
        Venta.total == total_venta,
        Venta.detalle_json == detalle_json_str,
        Venta.fecha >= tiempo_limite
    ).first()

    if venta_fantasma:
        print(f"✅ Eco masivo bloqueado (Patovica 60s): Venta de ${total_venta}")
        return jsonify({"ok": True, "mensaje": "Venta procesada (eco bloqueado)", "venta_id": venta_fantasma.id}), 200

    for item in items:
        p_id = item.get('producto_id') if item.get('producto_id') is not None else item.get('id')
        qty = item.get('qty', 0)
        if p_id and qty > 0:
            producto = Producto.query.get(int(p_id))
            if producto:
                producto.ventas_totales += qty
                if producto.stock >= qty:
                    producto.stock -= qty
                else:
                    producto.stock = 0

    venta_id = None
    if detalle or items:
        try:
            if not detalle:
                detalle = []
                for item in items:
                    p_id_fallback = item.get('producto_id') if item.get('producto_id') is not None else item.get('id')
                    p = Producto.query.get(int(p_id_fallback)) if p_id_fallback else None
                    if p:
                        precio_u = p.precio_lista_3 if lista_sel == 3 else (p.precio_lista_2 if lista_sel == 2 else p.precio)
                        detalle.append({
                            'producto_id': p.id,
                            'nombre': p.nombre,
                            'qty': item.get('qty', 1),
                            'precio_unit': precio_u
                        })
            
            # Soporte para medios de pago múltiples
            pagos = data.get('pagos', {})
            # monto_entregado_ef: lo que el cliente entrega físicamente (puede incluir vuelto)
            monto_entregado_ef = float(pagos.get('efectivo', 0))
            p_tr = float(pagos.get('transferencia', 0))
            p_db = float(pagos.get('debito', 0))
            p_cc = float(pagos.get('cc', 0))

            # Fallback para modo simple (un solo método)
            if not pagos and data.get('metodo_pago'):
                m = data.get('metodo_pago')
                if m == 'Efectivo': monto_entregado_ef = total_venta
                elif m in ['Mercado Pago', 'Transferencia']: p_tr = total_venta
                elif m == 'Débito': p_db = total_venta
                elif m == 'Cuenta Corriente': p_cc = total_venta

            # FIX CONTABLE (L-05): El monto real cobrado en efectivo es la venta menos lo pagado
            # por otros medios. El excedente entregado por el cliente (vuelto) NO debe sumarse a
            # los ingresos de la caja. Solo guardamos el total_venta como referencia de cobros reales.
            otros_medios = p_tr + p_db + p_cc
            # Lo que realmente se cobró en efectivo = total_venta - lo pagado por otros medios
            p_ef = max(0.0, round(min(monto_entregado_ef, total_venta - otros_medios), 2))

            # El monto entregado por el cliente se guarda como dato informativo (para calcular vuelto)
            monto_entregado_total = monto_entregado_ef + p_tr + p_db + p_cc
            vuelto = max(0.0, round(monto_entregado_total - total_venta, 2))

            if p_cc > 0 and cliente_id:
                cliente = db.session.get(Cliente, cliente_id)
                if cliente:
                    if cliente.limite_credito > 0 and (cliente.saldo + p_cc) > cliente.limite_credito:
                        db.session.rollback()
                        return jsonify({"ok": False, "mensaje": f"Límite de crédito excedido. Saldo: ${cliente.saldo:.2f}, Límite: ${cliente.limite_credito:.2f}"}), 403
                    cliente.saldo += p_cc

            venta = Venta(
                cliente_id=cliente_id,
                total=total_venta,
                detalle_json=_json.dumps(detalle, ensure_ascii=False),
                lista_precios=lista_sel,
                tipo=data.get('tipo', 'local'),
                metodo_pago=data.get('metodo_pago', 'Varios'),
                pago_efectivo=p_ef,          # Monto REAL cobrado en efectivo (sin vuelto)
                pago_transferencia=p_tr,
                pago_debito=p_db,
                pago_cc=p_cc,
                entregado=monto_entregado_ef, # Monto que el cliente entregó físicamente (informativo)
                fecha=hora_argentina(),
                sincronizado=not es_offline(),
                ultima_actualizacion=hora_argentina()
            )
            db.session.add(venta)
            db.session.flush()

            for d in detalle:
                prod_id = d.get('producto_id')
                if not prod_id:
                    for it in items:
                        pid_temp = it.get('producto_id') if it.get('producto_id') is not None else it.get('id')
                        if pid_temp:
                            p_temp = Producto.query.get(int(pid_temp))
                            if p_temp and p_temp.nombre == d['nombre']:
                                prod_id = p_temp.id
                                break
                
                det_obj = DetalleVenta(
                    venta_id=venta.id,
                    producto_id=prod_id,
                    nombre_producto=d['nombre'],
                    cantidad=d.get('qty', 1),
                    precio_unitario=d.get('precio_unit', 0.0),
                    descuento_porcentaje=d.get('discount_perc', 0.0),
                    subtotal=d.get('subtotal', 0.0)
                )
                db.session.add(det_obj)

            try:
                db.session.commit()
                venta_id = venta.id
            except IntegrityError as ie:
                db.session.rollback()
                print(f"⚠️ IntegrityError al hacer commit en registrar_venta: {ie}")
                return jsonify({"error": "Error de integridad de datos. Uno de los productos ya no existe o está inactivo. Vacía tu carrito y recarga el catálogo."}), 400

        except IntegrityError as ie:
            db.session.rollback()
            print(f"⚠️ IntegrityError capturado en registrar_venta: {ie}")
            return jsonify({"error": "Error de integridad de datos. Uno de los productos ya no existe o está inactivo. Vacía tu carrito y recarga el catálogo."}), 400
        except Exception as e:
            db.session.rollback()
            import traceback
            print(f"Error al registrar venta: {e}\n{traceback.format_exc()}")
            return jsonify({"ok": False, "mensaje": f"Error interno al guardar la venta: {str(e)}"}), 500
    else:
        try:
            db.session.commit()
        except IntegrityError as ie:
            db.session.rollback()
            print(f"⚠️ IntegrityError al hacer commit final: {ie}")
            return jsonify({"error": "Error de integridad de datos. Vacía tu carrito y recarga el catálogo."}), 400

    return jsonify({"ok": True, "mensaje": f"Venta registrada con éxito{mensaje_afip}", "venta_id": venta_id})

# ─── Rutas del Frontend (La Vidriera) ────────────────────────
# ─── Utilidades ────────────────────────────────────────────────────────
def to_title_case(text):
    if not text:
        return ""
    import re
    return re.sub(r"\b\w", lambda m: m.group(0).upper(), text.lower())

@app.route('/api/sincronizar', methods=['POST', 'GET'])
@login_requerido
def api_sincronizar():
    uri_nube = os.environ.get('DATABASE_URL')
    if not uri_nube:
        return jsonify({"ok": False, "mensaje": "DATABASE_URL no configurada en el servidor"}), 400
        
    clean_uri_nube = uri_nube
    if clean_uri_nube.startswith('postgres://'):
        clean_uri_nube = clean_uri_nube.replace('postgres://', 'postgresql://', 1)
        
    try:
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        
        # 1. Conectar a ambas bases de datos de forma explícita
        engine_local = create_engine(f'sqlite:///{DB_PATH}')
        SessionLocal = sessionmaker(bind=engine_local)
        session_local = SessionLocal()
        
        engine_nube = create_engine(clean_uri_nube, connect_args={'connect_timeout': 3})
        SessionNube = sessionmaker(bind=engine_nube)
        session_nube = SessionNube()
        
        # Verificar la conexión con Render
        session_nube.execute(text("SELECT 1"))
        
        # 2. Paso 1 (Subida): Productos modificados localmente a la nube
        productos_local_no_sinc = session_local.query(Producto).filter(
            (Producto.sincronizado == False) | (Producto.sincronizado == 0)
        ).all()
        
        for p_local in productos_local_no_sinc:
            p_nube = session_nube.query(Producto).filter(Producto.id == p_local.id).first()
            subir_cambio = True
            
            if p_nube:
                t_local = p_local.ultima_actualizacion.replace(tzinfo=None) if p_local.ultima_actualizacion else datetime.min
                t_nube = p_nube.ultima_actualizacion.replace(tzinfo=None) if p_nube.ultima_actualizacion else datetime.min
                
                if t_nube > t_local:
                    subir_cambio = False
                    print(f"[SYNC] Producto ID {p_local.id} '{p_local.nombre}' omitido para subir: Nube es más reciente.")
            
            if subir_cambio:
                if not p_nube:
                    p_nube = Producto(
                        id=p_local.id,
                        nombre=p_local.nombre,
                        descripcion=p_local.descripcion,
                        precio_lista_1=p_local.precio_lista_1,
                        precio_lista_2=p_local.precio_lista_2,
                        precio_lista_3=p_local.precio_lista_3,
                        precio_anterior=p_local.precio_anterior,
                        imagen=p_local.imagen,
                        imagen_url=p_local.imagen_url,
                        categoria_id=p_local.categoria_id,
                        stock=p_local.stock,
                        activo=p_local.activo,
                        favorito=p_local.favorito,
                        permitir_sin_stock=p_local.permitir_sin_stock,
                        ventas_totales=p_local.ventas_totales,
                        codigo_barra=p_local.codigo_barra,
                        descuento_volumen_activo=p_local.descuento_volumen_activo,
                        cantidad_minima_descuento=p_local.cantidad_minima_descuento,
                        porcentaje_descuento_volumen=p_local.porcentaje_descuento_volumen,
                        sincronizado=True,
                        ultima_actualizacion=p_local.ultima_actualizacion
                    )
                    session_nube.add(p_nube)
                else:
                    p_nube.nombre = p_local.nombre
                    p_nube.descripcion = p_local.descripcion
                    p_nube.precio_lista_1 = p_local.precio_lista_1
                    p_nube.precio_lista_2 = p_local.precio_lista_2
                    p_nube.precio_lista_3 = p_local.precio_lista_3
                    p_nube.precio_anterior = p_local.precio_anterior
                    p_nube.imagen = p_local.imagen
                    p_nube.imagen_url = p_local.imagen_url
                    p_nube.categoria_id = p_local.categoria_id
                    p_nube.stock = p_local.stock
                    p_nube.activo = p_local.activo
                    p_nube.favorito = p_local.favorito
                    p_nube.permitir_sin_stock = p_local.permitir_sin_stock
                    p_nube.ventas_totales = p_local.ventas_totales
                    p_nube.codigo_barra = p_local.codigo_barra
                    p_nube.descuento_volumen_activo = p_local.descuento_volumen_activo
                    p_nube.cantidad_minima_descuento = p_local.cantidad_minima_descuento
                    p_nube.porcentaje_descuento_volumen = p_local.porcentaje_descuento_volumen
                    p_nube.sincronizado = True
                    p_nube.ultima_actualizacion = p_local.ultima_actualizacion
            
            p_local.sincronizado = True
            
        session_nube.commit()
        session_local.commit()
        
        # Reiniciar secuencias de seriales de IDs en Postgres (Productos)
        try:
            session_nube.execute(text("SELECT setval(pg_get_serial_sequence('\"Productos\"', 'id'), COALESCE(MAX(id), 1)) FROM \"Productos\""))
            session_nube.commit()
        except Exception as seq_err:
            session_nube.rollback()
            print(f"[SYNC] No se pudo reiniciar secuencia de Productos: {seq_err}")
            
        # 3. Paso 1b (Subida): Ventas locales no sincronizadas a la nube
        ventas_local_no_sinc = session_local.query(Venta).filter(
            (Venta.sincronizado == False) | (Venta.sincronizado == 0)
        ).order_by(Venta.id.asc()).all()
        
        for v_local in ventas_local_no_sinc:
            v_nube = session_nube.query(Venta).filter(Venta.id == v_local.id).first()
            if not v_nube:
                v_nube = Venta(
                    id=v_local.id,
                    cliente_id=v_local.cliente_id,
                    fecha=v_local.fecha,
                    total=v_local.total,
                    detalle_json=v_local.detalle_json,
                    lista_precios=v_local.lista_precios,
                    tipo=v_local.tipo,
                    metodo_pago=v_local.metodo_pago,
                    pago_efectivo=v_local.pago_efectivo,
                    pago_transferencia=v_local.pago_transferencia,
                    pago_debito=v_local.pago_debito,
                    pago_cc=v_local.pago_cc,
                    entregado=v_local.entregado,
                    sincronizado=True,
                    ultima_actualizacion=v_local.ultima_actualizacion
                )
                session_nube.add(v_nube)
                
            v_local.sincronizado = True
            
        session_nube.commit()
        session_local.commit()
        
        # Reiniciar secuencias de seriales de IDs en Postgres (Ventas)
        try:
            session_nube.execute(text("SELECT setval(pg_get_serial_sequence('ventas', 'id'), COALESCE(MAX(id), 1)) FROM ventas"))
            session_nube.commit()
        except Exception as seq_err:
            session_nube.rollback()
            print(f"[SYNC] No se pudo reiniciar secuencia de Ventas: {seq_err}")
            
        # 4. Paso 2 (Bajada): Sincronizar categorías primero
        categorias_nube = session_nube.query(Categoria).all()
        for cat_nube in categorias_nube:
            cat_local = session_local.query(Categoria).filter(Categoria.id == cat_nube.id).first()
            if not cat_local:
                cat_local = Categoria(id=cat_nube.id, nombre=cat_nube.nombre)
                session_local.add(cat_local)
            else:
                cat_local.nombre = cat_nube.nombre
        session_local.commit()
        
        # Obtener el tiempo de la última sincronización local
        res = session_local.query(db.func.max(Producto.ultima_actualizacion)).filter(
            (Producto.sincronizado == True) | (Producto.sincronizado == 1)
        ).scalar()
        max_local_time = res if res else datetime.min
        
        # 5. Paso 2b (Bajada): Productos nuevos/actualizados en la nube a SQLite
        productos_nube_nuevos = session_nube.query(Producto).filter(
            Producto.ultima_actualizacion > max_local_time
        ).all()
        
        for p_nube in productos_nube_nuevos:
            p_local = session_local.query(Producto).filter(Producto.id == p_nube.id).first()
            t_nube = p_nube.ultima_actualizacion.replace(tzinfo=None) if p_nube.ultima_actualizacion else datetime.min
            actualizar_local = True
            
            if p_local:
                t_local = p_local.ultima_actualizacion.replace(tzinfo=None) if p_local.ultima_actualizacion else datetime.min
                if t_local > t_nube:
                    actualizar_local = False
                    print(f"[SYNC] Producto ID {p_nube.id} '{p_nube.nombre}' omitido para bajar: Local es más reciente.")
            
            if actualizar_local:
                if not p_local:
                    p_local = Producto(
                        id=p_nube.id,
                        nombre=p_nube.nombre,
                        descripcion=p_nube.descripcion,
                        precio_lista_1=p_nube.precio_lista_1,
                        precio_lista_2=p_nube.precio_lista_2,
                        precio_lista_3=p_nube.precio_lista_3,
                        precio_anterior=p_nube.precio_anterior,
                        imagen=p_nube.imagen,
                        imagen_url=p_nube.imagen_url,
                        categoria_id=p_nube.categoria_id,
                        stock=p_nube.stock,
                        activo=p_nube.activo,
                        favorito=p_nube.favorito,
                        permitir_sin_stock=p_nube.permitir_sin_stock,
                        ventas_totales=p_nube.ventas_totales,
                        codigo_barra=p_nube.codigo_barra,
                        descuento_volumen_activo=p_nube.descuento_volumen_activo,
                        cantidad_minima_descuento=p_nube.cantidad_minima_descuento,
                        porcentaje_descuento_volumen=p_nube.porcentaje_descuento_volumen,
                        sincronizado=True,
                        ultima_actualizacion=t_nube
                    )
                    session_local.add(p_local)
                else:
                    p_local.nombre = p_nube.nombre
                    p_local.descripcion = p_nube.descripcion
                    p_local.precio_lista_1 = p_nube.precio_lista_1
                    p_local.precio_lista_2 = p_nube.precio_lista_2
                    p_local.precio_lista_3 = p_nube.precio_lista_3
                    p_local.precio_anterior = p_nube.precio_anterior
                    p_local.imagen = p_nube.imagen
                    p_local.imagen_url = p_nube.imagen_url
                    p_local.categoria_id = p_nube.categoria_id
                    p_local.stock = p_nube.stock
                    p_local.activo = p_nube.activo
                    p_local.favorito = p_nube.favorito
                    p_local.permitir_sin_stock = p_nube.permitir_sin_stock
                    p_local.ventas_totales = p_nube.ventas_totales
                    p_local.codigo_barra = p_nube.codigo_barra
                    p_local.descuento_volumen_activo = p_nube.descuento_volumen_activo
                    p_local.cantidad_minima_descuento = p_nube.cantidad_minima_descuento
                    p_local.porcentaje_descuento_volumen = p_nube.porcentaje_descuento_volumen
                    p_local.sincronizado = True
                    p_local.ultima_actualizacion = t_nube
                    
        session_local.commit()
        return jsonify({"ok": True, "mensaje": "Sincronización bidireccional completada con éxito"}), 200
        
    except Exception as e:
        import traceback
        print(f"[SYNC FATAL ERROR] {e}\n{traceback.format_exc()}")
        return jsonify({"ok": False, "mensaje": f"Error de sincronización: {str(e)}"}), 500
    finally:
        if 'session_local' in locals():
            session_local.close()
        if 'session_nube' in locals():
            session_nube.close()

@app.route('/imprimir_ticket/')
@app.route('/imprimir_ticket/<int:id>')
@login_requerido # Protegido (A-05)
def endpoint_imprimir_ticket(id=None):
    if id is None:
        return "ID de ticket no proporcionado", 400
    
    venta = db.session.get(Venta, id)
    if not venta:
        abort(404, description="El comprobante no existe o fue anulado.")
        
    try:
        import json
        detalle = json.loads(venta.detalle_json or '[]')
        
        # Formatear nombres para el ticket (Title Case)
        for item in detalle:
            item['nombre'] = to_title_case(item.get('nombre', ''))
            
        # Regla de Privacidad: Título genérico y ocultar cliente si es CF
        es_venta_rapida = True
        cliente_nombre = ""
        if venta.cliente and venta.cliente.nombre != "Consumidor Final":
            cliente_nombre = to_title_case(venta.cliente.nombre)
            es_venta_rapida = False
            
        return render_template('ticket.html', 
                             venta=venta, 
                             detalle=detalle, 
                             cliente_nombre=cliente_nombre,
                             es_venta_rapida=es_venta_rapida)
    except Exception as e:
        return f"Error al generar el ticket: {str(e)}", 500



@app.before_request
def check_admin_auth():
    if request.method in ['POST', 'PUT', 'DELETE'] and request.path != '/':
        if not session.get('admin_autenticado'):
            if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"ok": False, "error": "No autorizado"}), 401
            return redirect('/')

    if request.path.startswith('/admin') or request.path.startswith('/facturador'):
        if not session.get('admin_autenticado'):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"ok": False, "error": "No autorizado"}), 401
            return redirect('/')

@app.route('/', methods=['GET', 'POST'])
def index():
    ip_cliente = request.remote_addr
    if session.get('admin_autenticado'):
        return redirect('/admin')

    if request.method == 'POST':
        registro = intentos_login.get(ip_cliente)
        if registro and registro['intentos'] >= 5:
            tiempo_transcurrido = time.time() - registro['ultimo_intento']
            if tiempo_transcurrido < 900:
                return render_template('index.html', error="Demasiados intentos fallidos. Por favor, espere 15 minutos."), 429

        username = request.form.get('username')
        password = request.form.get('password')
        usuario = Usuario.query.filter_by(username=username).first()
        if usuario and check_password_hash(usuario.password_hash, password):
            session['admin_autenticado'] = True
            session.permanent = True
            intentos_login.pop(ip_cliente, None)
            return redirect('/admin')
        else:
            registro = intentos_login.get(ip_cliente, {'intentos': 0, 'ultimo_intento': 0})
            registro['intentos'] += 1
            registro['ultimo_intento'] = time.time()
            intentos_login[ip_cliente] = registro
            return render_template('index.html', error='Credenciales incorrectas')
    return render_template('index.html')

@app.route('/facturador')
@login_requerido
def facturador():
    if not session.get('admin_autenticado'):
        return redirect('/')
    verificar_corte_automatico()
    return render_template('facturador.html')

@app.route('/ticket/<int:venta_id>')
@login_requerido # Protegido (A-05)
def endpoint_ticket_legacy(venta_id):
    return endpoint_imprimir_ticket(venta_id)

@app.route('/descargar_factura/<int:venta_id>')
@login_requerido # Protegido (A-05)
def descargar_factura(venta_id):
    venta = db.session.get(Venta, venta_id)
    if not venta:
        abort(404, description="Venta no encontrada")
        
    import json
    detalle = json.loads(venta.detalle_json or '[]')
    
    # Calcular altura dinámica: 70mm base + 10mm por item + 40mm pie
    item_count = len(detalle)
    page_height = 80 + (item_count * 8) + 40
    
    # FPDF(orientation, unit, format)
    # format=(ancho, alto) en mm
    pdf = FPDF('P', 'mm', (80, page_height))
    pdf.add_page()
    pdf.set_auto_page_break(False)
    
    # Encabezado
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 8, 'TODO GOLOSINA', 0, 1, 'C')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 4, 'Alberdi 950, Aguilares - Tucuman', 0, 1, 'C')
    pdf.cell(0, 4, '----------------------------------------------------------', 0, 1, 'C')
    
    pdf.ln(2)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, f'TICKET #000-{venta.id}', 0, 1, 'C')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 4, f"Fecha: {venta.fecha.strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
    pdf.ln(3)
    
    # Tabla de Productos
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(8, 6, 'Cant', 'B', 0, 'C')
    pdf.cell(37, 6, 'Producto', 'B', 0, 'L')
    pdf.cell(15, 6, 'Subtotal', 'B', 1, 'R')
    
    pdf.set_font('Arial', '', 8)
    for item in detalle:
        # Truncar nombre si es muy largo
        nombre = item.get('nombre', 'Prod')[:22]
        pdf.cell(8, 7, f"{item['qty']}", 0, 0, 'C')
        pdf.cell(37, 7, nombre, 0, 0, 'L')
        sub = item['qty'] * item['precio_unit']
        pdf.cell(15, 7, f"${sub:,.0f}".replace(',', '.'), 0, 1, 'R')
    
    pdf.ln(2)
    pdf.cell(0, 0, '', 'T', 1)
    pdf.ln(2)
    
    # Totales
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(35, 8, 'TOTAL:', 0, 0, 'L')
    pdf.cell(25, 8, f"${venta.total:,.2f}", 0, 1, 'R')
    
    pdf.ln(2)
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(0, 5, 'MEDIOS DE PAGO:', 0, 1, 'L')
    pdf.set_font('Arial', '', 8)
    if venta.pago_efectivo > 0: pdf.cell(0, 4, f"- Efectivo: ${venta.pago_efectivo:,.2f}", 0, 1, 'L')
    if venta.pago_transferencia > 0: pdf.cell(0, 4, f"- Transf/MP: ${venta.pago_transferencia:,.2f}", 0, 1, 'L')
    if venta.pago_debito > 0: pdf.cell(0, 4, f"- Débito: ${venta.pago_debito:,.2f}", 0, 1, 'L')
    if venta.pago_cc > 0: pdf.cell(0, 4, f"- A Cuenta: ${venta.pago_cc:,.2f}", 0, 1, 'L')

    if venta.pago_cc > 0 and venta.cliente:
        pdf.ln(1)
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(0, 5, f"Deuda Total Actual: ${venta.cliente.saldo:,.2f}", 0, 1, 'L')
    
    pdf.ln(5)
    pdf.set_font('Arial', 'I', 8)
    pdf.cell(0, 5, '¡Gracias por su compra!', 0, 1, 'C')
    pdf.cell(0, 5, 'www.todogolosina.com', 0, 1, 'C')

    # Generar salida
    output = io.BytesIO()
    # fpdf.output returns the pdf content as a string (py2) or bytes (py3) in dest='S'
    pdf_content = pdf.output(dest='S')
    if isinstance(pdf_content, str):
        pdf_content = pdf_content.encode('latin-1')
    
    output.write(pdf_content)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        download_name=f'Ticket_TodoGolosina_{venta_id}.pdf',
        as_attachment=True
    )

@app.route('/logout')
def logout():
    session.pop('admin_autenticado', None)
    return redirect('/')

@app.route('/api/cambiar-password', methods=['POST'])
@login_requerido
def cambiar_password():
    if not session.get('admin_autenticado'):
        return jsonify({'ok': False, 'mensaje': 'No autenticado'}), 401
    data = request.get_json()
    if not data or not data.get('nueva_password'):
        return jsonify({'ok': False, 'mensaje': 'Debe ingresar una nueva contraseña'}), 400
    nueva = data['nueva_password'].strip()
    if len(nueva) < 6:
        return jsonify({'ok': False, 'mensaje': 'La contraseña debe tener al menos 6 caracteres'}), 400
    usuario = Usuario.query.filter_by(username='admin').first()
    if not usuario:
        return jsonify({'ok': False, 'mensaje': 'Usuario admin no encontrado'}), 404
    usuario.password_hash = generate_password_hash(nueva)
    db.session.commit()
    return jsonify({'ok': True, 'mensaje': 'Contraseña actualizada correctamente'})

@app.route('/api/clientes/<int:id>', methods=['DELETE'])
@login_requerido
def delete_cliente(id):
    try:
        cliente = Cliente.query.get(id)
        if not cliente:
            return jsonify({"ok": False, "mensaje": "Cliente no encontrado"}), 404
        db.session.delete(cliente)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        print(f"Error al eliminar cliente: {e}")
        return jsonify({'error': str(e)}), 500

# ─── API Clientes ────────────────────────────────────────────────────────────────
@app.route('/api/clientes', methods=['GET'])
@app.route('/buscar_clientes', methods=['GET'])
@app.route('/obtener_clientes', methods=['GET'])
@login_requerido
def get_clientes():
    try:
        buscar = request.args.get('q', '').strip()
        query = Cliente.query.filter_by(activo=1)
        if buscar:
            query = query.filter(Cliente.nombre.ilike(f'%{buscar}%'))
        clientes = query.order_by(Cliente.nombre.asc()).all()
        return jsonify({"ok": True, "clientes": [c.to_dict() for c in clientes]})
    except Exception as e:
        return jsonify({"error_interno": str(e), "detalle": traceback.format_exc()}), 500

@app.route('/api/clientes', methods=['POST'])
@app.route('/guardar_cliente', methods=['POST'])
@login_requerido
def add_cliente():
    try:
        data = request.json or {}
        nombre = data.get('nombre', '').strip()
        cuit = data.get('cuit', '').strip()
        if not nombre:
            return jsonify({"ok": False, "mensaje": "El nombre es obligatorio"}), 400
        
        # Validación de Duplicados (DNI/CUIT)
        if cuit:
            existente = Cliente.query.filter_by(cuit=cuit).first()
            if existente:
                return jsonify({"ok": False, "mensaje": "Este cliente ya está registrado con ese DNI/CUIT"}), 400
        else:
            # Si no hay CUIT, validamos por nombre exacto para evitar duplicados basura
            existente_nombre = Cliente.query.filter(Cliente.nombre.ilike(nombre)).first()
            if existente_nombre:
                return jsonify({"ok": False, "mensaje": f"Ya existe un cliente registrado con el nombre '{nombre}'"}), 400

        cliente = Cliente(
            nombre=nombre,
            cuit=cuit,
            condicion_iva=data.get('condicion_iva', 'Consumidor Final').strip(),
            telefono=data.get('telefono', '').strip(),
            direccion=data.get('direccion', '').strip(),
            descuento_fijo=float(data.get('descuento_fijo', 0.0))
        )
        db.session.add(cliente)
        db.session.commit()
        return jsonify({"ok": True, "cliente": cliente.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error_interno": str(e), "detalle": traceback.format_exc()}), 500

@app.route('/api/clientes/<int:id>', methods=['PUT'])
@app.route('/editar_cliente/<int:id>', methods=['POST', 'PUT'])
@login_requerido
def edit_cliente(id):
    cliente = db.session.get(Cliente, id)
    if not cliente:
        return jsonify({"ok": False, "mensaje": "Cliente no encontrado"}), 404

    data = request.json or {}
    nombre = data.get('nombre', '').strip()
    cuit   = data.get('cuit', '').strip()

    if not nombre:
        return jsonify({"ok": False, "mensaje": "El nombre es obligatorio"}), 400

    # ── Validación inteligente: excluye al cliente actual ──
    if cuit:
        duplicado = Cliente.query.filter(
            Cliente.cuit == cuit,
            Cliente.id != id
        ).first()
        if duplicado:
            return jsonify({"ok": False, "mensaje": f"Ya existe otro cliente con ese DNI/CUIT ({duplicado.nombre})"}), 400

    try:
        # ── Actualizar todos los campos ──
        cliente.nombre        = nombre
        cliente.cuit          = cuit
        cliente.condicion_iva = data.get('condicion_iva', cliente.condicion_iva or 'Consumidor Final').strip()
        cliente.telefono      = data.get('telefono', '').strip()
        cliente.direccion     = data.get('direccion', '').strip()
        cliente.descuento_fijo = float(data.get('descuento_fijo', cliente.descuento_fijo or 0.0))

        db.session.commit()
        return jsonify({"ok": True, "cliente": cliente.to_dict()})
    except Exception as e:
        db.session.rollback()
        print(f"ERROR AL EDITAR CLIENTE: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/clientes/deudores', methods=['GET'])
@login_requerido
def get_clientes_deudores():
    try:
        clientes = Cliente.query.filter(Cliente.saldo > 0, Cliente.activo == 1).order_by(Cliente.saldo.desc()).all()
        return jsonify({"ok": True, "clientes": [c.to_dict() for c in clientes]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clientes/registrar_pago', methods=['POST'])
@login_requerido
def registrar_pago():
    data = request.json or {}
    cliente_id = data.get('cliente_id')
    monto = float(data.get('monto', 0))
    if not cliente_id or monto <= 0:
        return jsonify({"ok": False, "mensaje": "Datos inválidos"}), 400
    
    cliente = db.session.get(Cliente, cliente_id)
    if not cliente:
        return jsonify({"ok": False, "mensaje": "Cliente no encontrado"}), 404
        
    try:
        cliente.saldo -= monto
        # Registrar como ingreso en caja
        pago_mov = Gasto(
            descripcion=f"Cobranza: {cliente.nombre}",
            monto=monto,
            categoria="Cobranza",
            tipo="Ingreso"
        )
        db.session.add(pago_mov)
        db.session.commit()
        return jsonify({"ok": True, "nuevo_saldo": cliente.saldo})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/caja/estado', methods=['GET'])
@login_requerido
def get_caja_estado():
    try:
        caja = CajaDiaria.query.filter_by(estado='Abierta').order_by(CajaDiaria.id.desc()).first()
        if caja:
            return jsonify({
                "ok": True,
                "abierta": True,
                "monto_inicial": caja.monto_inicial,
                "caja": caja.to_dict()
            }), 200
        return jsonify({
            "ok": True,
            "abierta": False
        }), 200
    except Exception as e:
        return jsonify({"ok": False, "abierta": False, "error": str(e)}), 500

@app.route('/api/caja/abrir', methods=['POST'])
@login_requerido
def abrir_caja():
    data = request.json or {}
    monto = float(data.get('monto_inicial', 0))
    # Cerrar cualquier caja que haya quedado abierta por error antes de abrir una nueva?
    # O simplemente no permitir abrir si ya hay una.
    existente = CajaDiaria.query.filter_by(estado='Abierta').first()
    if existente:
        return jsonify({"ok": False, "mensaje": "Ya existe una caja abierta"}), 400
    
    try:
        tz = pytz.timezone('America/Argentina/Buenos_Aires')
        ahora = datetime.now(tz)
        turno_asignado = 'Mañana' if ahora.hour < 13 else 'Tarde'

        nueva = CajaDiaria(monto_inicial=monto, estado='Abierta', turno=turno_asignado)
        db.session.add(nueva)
        db.session.commit()
        return jsonify({"ok": True, "caja": nueva.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "mensaje": str(e)}), 500

@app.route('/api/caja/cerrar', methods=['POST'])
@login_requerido
def cerrar_caja():
    caja = CajaDiaria.query.filter_by(estado='Abierta').order_by(CajaDiaria.id.desc()).first()
    if not caja:
        return jsonify({"ok": False, "mensaje": "No hay una caja abierta para cerrar"}), 400
    
    try:
        # Aquí se podrían calcular totales finales y guardarlos si el modelo tuviera esos campos
        caja.estado = 'Cerrada'
        caja.fecha_cierre = hora_argentina()
        db.session.commit()
        return jsonify({"ok": True, "mensaje": "Caja cerrada correctamente"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "mensaje": str(e)}), 500

@app.route('/api/ventas_hoy', methods=['GET'])
@login_requerido
def get_ventas_hoy():
    verificar_corte_automatico()
    try:
        fecha_inicio_str = request.args.get('inicio')
        fecha_fin_str = request.args.get('fin')
        
        caja = CajaDiaria.query.filter_by(estado='Abierta').order_by(CajaDiaria.id.desc()).first()

        if fecha_inicio_str and fecha_fin_str:
            inicio_rango = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fin_rango = datetime.strptime(fecha_fin_str + " 23:59:59", '%Y-%m-%d %H:%M:%S')
            ventas = Venta.query.filter(Venta.fecha >= inicio_rango, Venta.fecha <= fin_rango).order_by(Venta.fecha.desc()).all()
            gastos = Gasto.query.filter(Gasto.fecha >= inicio_rango, Gasto.fecha <= fin_rango).all()
        else:
            if caja:
                inicio_rango = caja.fecha_apertura
            else:
                inicio_rango = datetime.combine(date.today(), time.min)

            ventas = Venta.query.filter(Venta.fecha >= inicio_rango).order_by(Venta.fecha.desc()).all()
            gastos = Gasto.query.filter(Gasto.fecha >= inicio_rango).all()
        
        m_ef = sum((v.pago_efectivo or 0.0) for v in ventas)
        m_tr = sum((v.pago_transferencia or 0.0) for v in ventas)
        m_db = sum((v.pago_debito or 0.0) for v in ventas)
        m_cc = sum((v.pago_cc or 0.0) for v in ventas)
        
        total_ventas = sum((v.total or 0.0) for v in ventas)
        
        ingresos_extra = sum((g.monto or 0.0) for g in gastos if g.tipo == 'Ingreso' and g.categoria != 'Cobranza')
        cobranzas = sum((g.monto or 0.0) for g in gastos if g.tipo == 'Ingreso' and g.categoria == 'Cobranza')
        egresos = sum((g.monto or 0.0) for g in gastos if g.tipo == 'Egreso')
        
        monto_inicial = caja.monto_inicial if caja else 0.0
        # Efectivo Real = Inicio + Ventas Efectivo + Cobranzas + Ingresos Extra - Egresos
        efectivo_real = monto_inicial + m_ef + cobranzas + ingresos_extra - egresos

        # Métricas extra para el Dashboard
        tickets_hoy = len(ventas)
        clientes_hoy = len(set(v.cliente_id for v in ventas if v.cliente_id))
        alertas_stock = Producto.query.filter(Producto.activo == 1, Producto.stock <= 5).count()

        return jsonify({
            "ok": True,
            "caja_abierta": True if caja else False,
            "monto_inicial": monto_inicial,
            "total_ventas": total_ventas,
            "tickets_hoy": tickets_hoy,
            "clientes_hoy": clientes_hoy,
            "alertas_stock": alertas_stock,
            "metodos": {
                "efectivo": m_ef,
                "transferencia": m_tr,
                "debito": m_db,
                "cc": m_cc
            },
            "ingresos_extra": ingresos_extra,
            "cobranzas": cobranzas,
            "egresos": egresos,
            "efectivo_real": efectivo_real,
            "ventas": [{
                "id": v.id,
                "hora": v.fecha.strftime('%H:%M') if hasattr(v.fecha, 'strftime') else (str(v.fecha)[:16] if v.fecha else '--:--'),
                "metodo_pago": getattr(v, 'metodo_pago', 'No especificado') or 'No especificado',
                "total": v.total or 0.0
            } for v in ventas],
            "gastos": [g.to_dict() for g in gastos]
        })
    except Exception as e:
        print(f"Error en get_ventas_hoy: {e}")
        return jsonify([])

@app.route('/api/venta/<int:id>', methods=['GET'])
@login_requerido
def get_venta_detalle(id):
    try:
        venta = db.session.get(Venta, id)
        if not venta:
            return jsonify({"ok": False, "mensaje": "Venta no encontrada"}), 404
        
        import json
        detalle = json.loads(venta.detalle_json or '[]')
        
        for item in detalle:
            item['nombre'] = to_title_case(item.get('nombre', ''))
            
        cliente_nombre = "Consumidor Final"
        if venta.cliente:
            cliente_nombre = to_title_case(venta.cliente.nombre)
            
        return jsonify({
            "ok": True,
            "id": venta.id,
            "fecha": venta.fecha.strftime('%d/%m/%Y %H:%M:%S') if hasattr(venta.fecha, 'strftime') else str(venta.fecha),
            "metodo_pago": getattr(venta, 'metodo_pago', 'No especificado') or 'No especificado',
            "total": venta.total or 0.0,
            "cliente_nombre": cliente_nombre,
            "detalle": detalle
        }), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/ventas/<int:id_venta>', methods=['GET'])
@login_requerido
def obtener_detalle_venta(id_venta):
    venta = db.session.get(Venta, id_venta)
    if not venta:
        return jsonify({"ok": False, "mensaje": "El comprobante no existe o fue anulado."}), 404
    
    items_vendidos = []
    
    # 1. Intentar obtener de la relación de base de datos si existe
    if hasattr(venta, 'detalles') and venta.detalles: 
        for detalle in venta.detalles:
            items_vendidos.append({
                "cantidad": detalle.cantidad,
                "nombre": to_title_case(detalle.producto.nombre if detalle.producto else "Producto Eliminado"),
                "precio": detalle.precio_unitario,
                "subtotal": detalle.cantidad * detalle.precio_unitario
            })
            
    # 2. Si no hay relación o está vacía, deserializar desde detalle_json
    if not items_vendidos and hasattr(venta, 'detalle_json') and venta.detalle_json:
        try:
            import json
            detalle = json.loads(venta.detalle_json or '[]')
            for item in detalle:
                qty = item.get('qty') or item.get('cantidad') or 1
                precio = item.get('precio_unit') or item.get('precio_unitario') or item.get('precio') or 0.0
                nombre = to_title_case(item.get('nombre') or "Producto")
                items_vendidos.append({
                    "cantidad": qty,
                    "nombre": nombre,
                    "precio": precio,
                    "subtotal": qty * precio
                })
        except Exception:
            pass
            
    return jsonify({
        "id": venta.id,
        "fecha": venta.fecha.strftime('%d/%m/%Y %H:%M') if (venta.fecha and hasattr(venta.fecha, 'strftime')) else "",
        "total": venta.total,
        "items": items_vendidos
    }), 200

@app.route('/api/gastos', methods=['POST'])
@login_requerido
def add_gasto():
    data = request.json or {}
    descripcion = data.get('descripcion', '').strip()
    monto = data.get('monto', 0)
    if not descripcion or not monto:
        return jsonify({"ok": False, "mensaje": "Descripción y monto son obligatorios"}), 400
    
    from datetime import timedelta
    tiempo_limite = hora_argentina() - timedelta(seconds=15)
    
    # Buscar si hay un movimiento idéntico reciente
    mov_fantasma = Gasto.query.filter(
        Gasto.monto == float(monto),
        Gasto.descripcion == descripcion,
        Gasto.tipo == data.get('tipo', 'Egreso'),
        Gasto.fecha >= tiempo_limite
    ).first()
    
    if mov_fantasma:
        return jsonify({"ok": True, "mensaje": "Movimiento procesado (eco bloqueado)"}), 200
    
    try:
        gasto = Gasto(
            descripcion=descripcion, 
            monto=float(monto), 
            categoria=data.get('categoria', 'General'),
            tipo=data.get('tipo', 'Egreso')
        )
        db.session.add(gasto)
        db.session.commit()
        return jsonify({"ok": True, "gasto": gasto.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "mensaje": str(e)}), 500

@app.route('/api/ventas', methods=['GET'])
@login_requerido
def get_ventas_general():
    import json as _json
    buscar = request.args.get('q', '').strip()
    query = Venta.query.outerjoin(Cliente, Venta.cliente_id == Cliente.id)
    if buscar:
        query = query.filter(Cliente.nombre.ilike(f'%{buscar}%'))
    
    ventas = query.order_by(Venta.fecha.desc()).limit(100).all()
    
    return jsonify({
        "ok": True,
        "ventas": [{
            "id": v.id,
            "fecha": v.fecha.strftime('%d/%m/%Y %H:%M') if v.fecha else '',
            "cliente_nombre": v.cliente.nombre if v.cliente else 'N/A',
            "total": v.total,
            "detalle": _json.loads(v.detalle_json or '[]')
        } for v in ventas]
    })

@app.route('/api/ventas-cliente/<int:cliente_id>', methods=['GET'])
@login_requerido
def get_ventas_cliente(cliente_id):
    import json as _json
    cliente = db.session.get(Cliente, cliente_id)
    if not cliente:
        return jsonify({"ok": False, "mensaje": "Cliente no encontrado"}), 404
    ventas = Venta.query.filter_by(cliente_id=cliente_id).order_by(Venta.fecha.desc()).limit(20).all()
    return jsonify({
        "ok": True,
        "cliente": cliente.to_dict(),
        "ventas": [{
            "id": v.id,
            "fecha": v.fecha.strftime('%d/%m/%Y %H:%M') if v.fecha else '',
            "total": v.total,
            "detalle": _json.loads(v.detalle_json or '[]')
        } for v in ventas]
    })


# ─── Panel de Administración ──────────────────────────────────

@app.route('/api/buscar_productos', methods=['GET'])
@login_requerido
def api_buscar_productos():
    if not session.get('admin_autenticado') and not session.get('facturador_auth'):
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"ok": True, "productos": []})
        
    productos = Producto.query.filter(
        Producto.activo == 1,
        (Producto.nombre.ilike(f'%{q}%')) | (Producto.codigo_barra.ilike(f'%{q}%'))
    ).order_by(Producto.nombre.asc()).limit(30).all()
    
    return jsonify({
        "ok": True,
        "productos": [
            {
                "id": p.id,
                "nombre": p.nombre,
                "codigo_barra": p.codigo_barra,
                "precio_lista_1": p.precio_lista_1,
                "stock": p.stock,
                "categoria_nombre": p.categoria_rel.nombre if p.categoria_rel else "S/C"
            } for p in productos
        ]
    })


@app.route('/admin')
@login_requerido
def admin_dashboard():
    if not session.get('admin_autenticado'): return redirect('/')
    try:
        q    = request.args.get('q', '').strip()
        page = request.args.get('page', 1, type=int)
        base = Producto.query.filter_by(activo=1)
        if q:
            base = base.filter(
                or_(
                    Producto.nombre.ilike(f'%{q}%'),
                    Producto.codigo_barra.ilike(f'%{q}%')
                )
            )
        productos  = base.order_by(Producto.id.desc()).paginate(page=page, per_page=50, error_out=False)
        categorias = Categoria.query.all()
        return render_template('admin.html', productos=productos, categorias=categorias, search=q, q=q)
    except Exception as e:
        return f"<h1>Error Oculto: {str(e)}</h1><pre>{traceback.format_exc()}</pre>", 500

@app.route('/admin/producto/add', methods=['POST'])
@login_requerido
def admin_add_product():
    if not session.get('admin_autenticado'): return redirect('/')
    nombre = request.form.get('nombre')
    
    precio_str = request.form.get('precio', '0').strip().replace(',', '.')
    try:
        precio = float(precio_str) if precio_str else 0.0
    except ValueError:
        precio = 0.0
        
    precio_ant_str = request.form.get('precio_anterior', '').strip().replace(',', '.')
    try:
        precio_anterior = float(precio_ant_str) if precio_ant_str else None
    except ValueError:
        precio_anterior = None
        
    descripcion = request.form.get('descripcion', '')
    imagen_url = request.form.get('imagen_url', '')
    categoria_id_str = request.form.get('categoria_id')
    categoria_id = int(categoria_id_str) if categoria_id_str and categoria_id_str.isdigit() else None
    
    stock_str = request.form.get('stock', '0').strip()
    try:
        stock = int(stock_str) if stock_str else 0
    except ValueError:
        stock = 0

    favorito = True if request.form.get('favorito') else False
    permitir_sin_stock = True if request.form.get('permitir_sin_stock') else False

    # Triple Lista de Precios
    precio_1_str = request.form.get('precio', '0').strip().replace(',', '.')
    precio_2_str = request.form.get('precio_2', '').strip().replace(',', '.')
    precio_3_str = request.form.get('precio_3', '').strip().replace(',', '.')
    
    try:
        precio_lista_1 = float(precio_1_str) if precio_1_str else 0.0
    except ValueError:
        precio_lista_1 = 0.0

    try:
        precio_lista_2 = float(precio_2_str) if precio_2_str else precio_lista_1
    except ValueError:
        precio_lista_2 = precio_lista_1
    try:
        precio_lista_3 = float(precio_3_str) if precio_3_str else precio_lista_1
    except ValueError:
        precio_lista_3 = precio_lista_1

    # Descuento por volumen
    descuento_volumen_activo = True if request.form.get('descuento_volumen_activo') else False
    cant_min_str = request.form.get('cantidad_minima_descuento', '').strip()
    porc_desc_str = request.form.get('porcentaje_descuento_volumen', '').strip()
    try:
        cantidad_minima_descuento = int(cant_min_str) if cant_min_str else None
    except ValueError:
        cantidad_minima_descuento = None
    try:
        porcentaje_descuento_volumen = float(porc_desc_str) if porc_desc_str else None
    except ValueError:
        porcentaje_descuento_volumen = None

    codigo_barra = request.form.get('codigo_barra', '').strip()
    if codigo_barra:
        # Validación multi-variante: separar por comas y verificar cada uno
        nuevos_codigos = [c.strip() for c in codigo_barra.split(',') if c.strip()]
        filtros = [Producto.codigo_barra.ilike(f'%{c}%') for c in nuevos_codigos]
        posibles_duplicados = Producto.query.filter(or_(*filtros), Producto.activo == 1).all()

        for p in posibles_duplicados:
            codigos_existentes = [c.strip() for c in p.codigo_barra.split(',') if c.strip()]
            for nc in nuevos_codigos:
                if nc in codigos_existentes:
                     return jsonify({"error": f"El código '{nc}' ya está registrado en el producto: {p.nombre}"}), 400

    nuevo = Producto(
        nombre=nombre, precio_lista_1=precio_lista_1, precio_lista_2=precio_lista_2, precio_lista_3=precio_lista_3,
        precio_anterior=precio_anterior, descripcion=descripcion,
        imagen_url=imagen_url, categoria_id=categoria_id, stock=stock,
        favorito=favorito, permitir_sin_stock=permitir_sin_stock,
        codigo_barra=codigo_barra,
        descuento_volumen_activo=descuento_volumen_activo,
        cantidad_minima_descuento=cantidad_minima_descuento,
        porcentaje_descuento_volumen=porcentaje_descuento_volumen,
        sincronizado=not es_offline(),
        ultima_actualizacion=hora_argentina()
    )
    db.session.add(nuevo)
    db.session.commit()
    actualizar_version_catalogo()
    flash('Producto agregado exitosamente.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/producto/edit/<int:id>', methods=['POST'])
@login_requerido
def admin_edit_product(id):
    if not session.get('admin_autenticado'): return redirect('/')
    producto = db.session.get(Producto, id)
    if not producto:
        flash('Producto no encontrado.', 'danger')
        return redirect(url_for('admin_dashboard'))

    producto.nombre = request.form.get('nombre')
        
    precio_ant_str = request.form.get('precio_anterior', '').strip().replace(',', '.')
    try:
        producto.precio_anterior = float(precio_ant_str) if precio_ant_str else None
    except ValueError:
        producto.precio_anterior = None
        
    producto.descripcion = request.form.get('descripcion', '')
    producto.imagen_url = request.form.get('imagen_url', '')
    
    cat_id_str = request.form.get('categoria_id')
    producto.categoria_id = int(cat_id_str) if cat_id_str and cat_id_str.isdigit() else None
    
    stock_str = request.form.get('stock', '0').strip()
    try:
        producto.stock = int(stock_str) if stock_str else 0
    except ValueError:
        producto.stock = 0

    producto.favorito = True if request.form.get('favorito') else False
    producto.permitir_sin_stock = True if request.form.get('permitir_sin_stock') else False
    producto.codigo_barra = request.form.get('codigo_barra', '').strip()

    # Descuento por volumen
    producto.descuento_volumen_activo = True if request.form.get('descuento_volumen_activo') else False
    cant_min_str = request.form.get('cantidad_minima_descuento', '').strip()
    porc_desc_str = request.form.get('porcentaje_descuento_volumen', '').strip()
    try:
        producto.cantidad_minima_descuento = int(cant_min_str) if cant_min_str else None
    except ValueError:
        producto.cantidad_minima_descuento = None
    try:
        producto.porcentaje_descuento_volumen = float(porc_desc_str) if porc_desc_str else None
    except ValueError:
        producto.porcentaje_descuento_volumen = None

    # Validación de Duplicados en Edición (Multi-variante)
    nuevo_codigo_str = request.form.get('codigo_barra', '').strip()
    if nuevo_codigo_str:
        nuevos_codigos = [c.strip() for c in nuevo_codigo_str.split(',') if c.strip()]
        filtros = [Producto.codigo_barra.ilike(f'%{c}%') for c in nuevos_codigos]
        # Buscamos en OTROS productos
        posibles_duplicados = Producto.query.filter(or_(*filtros), Producto.id != id, Producto.activo == 1).all()

        for p in posibles_duplicados:
            codigos_existentes = [c.strip() for c in p.codigo_barra.split(',') if c.strip()]
            for nc in nuevos_codigos:
                if nc in codigos_existentes:
                     return jsonify({"error": f"El código '{nc}' ya está registrado en el producto: {p.nombre}"}), 400

    try:
        # Triple Lista de Precios
        p1_str = request.form.get('precio', '0').strip().replace(',', '.')
        p2_str = request.form.get('precio_2', '').strip().replace(',', '.')
        p3_str = request.form.get('precio_3', '').strip().replace(',', '.')
        try:
            producto.precio_lista_1 = float(p1_str) if p1_str else 0.0
        except ValueError:
            producto.precio_lista_1 = 0.0
        try:
            producto.precio_lista_2 = float(p2_str) if p2_str else producto.precio_lista_1
        except ValueError:
            producto.precio_lista_2 = producto.precio_lista_1
        try:
            producto.precio_lista_3 = float(p3_str) if p3_str else producto.precio_lista_1
        except ValueError:
            producto.precio_lista_3 = producto.precio_lista_1

        producto.sincronizado = not es_offline()
        producto.ultima_actualizacion = hora_argentina()
        db.session.commit()
        actualizar_version_catalogo()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"ok": True, "mensaje": "Producto editado correctamente."})
            
        flash('Producto editado correctamente.', 'info')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"error": f"Error de integridad: {str(e)}"}), 500
        flash(f'Error al editar producto: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/producto/delete/<int:id>', methods=['POST'])
@login_requerido
def admin_delete_product(id):
    if not session.get('admin_autenticado'): return redirect('/')
    try:
        producto = Producto.query.get_or_404(id)
        db.session.delete(producto)
        db.session.commit()
        actualizar_version_catalogo()
        flash('Producto eliminado definitivamente de la base de datos.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al intentar eliminar el producto (puede tener ventas asociadas): {str(e)}', 'danger')
    return redirect(url_for('admin_dashboard'))

@app.route('/api/productos/eliminar_masivo', methods=['POST'])
@login_requerido
def eliminar_masivo():
    data = request.json or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'ok': False, 'mensaje': 'No se seleccionaron productos'}), 400
    
    try:
        for p_id in ids:
            producto = db.session.get(Producto, p_id)
            if producto:
                producto.activo = 0
                producto.sincronizado = not es_offline()
                producto.ultima_actualizacion = hora_argentina()
        db.session.commit()
        actualizar_version_catalogo()
        return jsonify({'ok': True, 'mensaje': f'{len(ids)} productos eliminados correctamente.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'mensaje': f'Error al eliminar masivamente: {str(e)}'}), 500
@app.route('/api/productos/vaciar_catalogo', methods=['POST'])
@login_requerido
def vaciar_catalogo():
    if not session.get('admin_autenticado'):
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401
    try:
        # Borrado lógico masivo
        Producto.query.filter_by(activo=1).update({'activo': 0})
        db.session.commit()
        actualizar_version_catalogo()
        return jsonify({'ok': True, 'mensaje': 'Catálogo vaciado correctamente.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'mensaje': f'Error al vaciar catálogo: {str(e)}'}), 500

@app.route('/api/sincronizar-bd', methods=['POST'])
@login_requerido
def sincronizar_bd():
    if not session.get('admin_autenticado'):
        return jsonify({'ok': False, 'mensaje': 'No autorizado'}), 401
    
    data = request.json or {}
    ids_validos = data.get('ids_validos')
    
    if not isinstance(ids_validos, list):
        return jsonify({'ok': False, 'mensaje': 'Formato de ids_validos incorrecto, debe ser una lista.'}), 400

    try:
        if not ids_validos:
            productos_sobrantes = Producto.query.all()
        else:
            productos_sobrantes = Producto.query.filter(~Producto.id.in_(ids_validos)).all()
            
        eliminados = 0
        desactivados = 0
        
        for p in productos_sobrantes:
            tiene_ventas = DetalleVenta.query.filter_by(producto_id=p.id).first() is not None
            if tiene_ventas:
                if p.activo != 0:
                    p.activo = 0
                    p.sincronizado = not es_offline()
                    p.ultima_actualizacion = hora_argentina()
                    desactivados += 1
            else:
                db.session.delete(p)
                eliminados += 1
                
        db.session.commit()
        try:
            actualizar_version_catalogo()
        except:
            pass # Para evitar error si no existe la funcion
        return jsonify({
            'ok': True, 
            'mensaje': f'Sincronización exitosa: {eliminados} eliminados, {desactivados} desactivados.'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'mensaje': f'Error al sincronizar BD: {str(e)}'}), 500

@app.route('/api/productos/aumento_masivo', methods=['POST'])
@login_requerido
def aumento_masivo():
    data = request.json or {}
    ids = data.get('ids', [])
    try:
        porcentaje = float(data.get('porcentaje', 0))
    except ValueError:
        return jsonify({'ok': False, 'mensaje': 'Porcentaje inválido.'}), 400
        
    if not ids:
        return jsonify({'ok': False, 'mensaje': 'No se seleccionaron productos'}), 400
    
    try:
        factor = 1.0 + (porcentaje / 100.0)
        for p_id in ids:
            producto = db.session.get(Producto, p_id)
            if producto:
                producto.precio_anterior = producto.precio_lista_1
                producto.precio_lista_1 = round(producto.precio_lista_1 * factor, 2)
                if producto.precio_lista_2:
                    producto.precio_lista_2 = round(producto.precio_lista_2 * factor, 2)
                if producto.precio_lista_3:
                    producto.precio_lista_3 = round(producto.precio_lista_3 * factor, 2)
                producto.sincronizado = not es_offline()
                producto.ultima_actualizacion = hora_argentina()
        db.session.commit()
        actualizar_version_catalogo()
        return jsonify({'ok': True, 'mensaje': f'Precios de {len(ids)} productos aumentados un {porcentaje}% correctamente.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'mensaje': f'Error al aumentar precios masivamente: {str(e)}'}), 500

# ─── CRUD de Categorías ──────────────────────────────────────
@app.route('/admin/categorias')
@login_requerido
def admin_categorias():
    if not session.get('admin_autenticado'): return redirect('/')
    categorias = Categoria.query.all()
    return render_template('admin_categorias.html', categorias=categorias)

@app.route('/admin/categoria/add', methods=['POST'])
@login_requerido
def admin_add_categoria():
    if not session.get('admin_autenticado'): return redirect('/')
    nombre = request.form.get('nombre')
    if nombre:
        nueva = Categoria(nombre=nombre)
        db.session.add(nueva)
        try:
            db.session.commit()
            flash('Categoría agregada exitosamente.', 'success')
        except:
            db.session.rollback()
            flash('Error: La categoría ya existe u ocurrió un error.', 'danger')
    return redirect(url_for('admin_categorias'))

@app.route('/admin/categoria/edit/<int:id>', methods=['POST'])
@login_requerido
def admin_edit_categoria(id):
    if not session.get('admin_autenticado'): return redirect('/')
    categoria = db.session.get(Categoria, id)
    nombre = request.form.get('nombre')
    if categoria and nombre:
        categoria.nombre = nombre
        try:
            db.session.commit()
            flash('Categoría editada exitosamente.', 'info')
        except:
            db.session.rollback()
            flash('Error: La categoría ya existe.', 'danger')
    return redirect(url_for('admin_categorias'))

@app.route('/admin/categoria/delete/<int:id>', methods=['POST'])
@login_requerido
def admin_delete_categoria(id):
    if not session.get('admin_autenticado'): return redirect('/')
    categoria = db.session.get(Categoria, id)
    if categoria:
        if Producto.query.filter_by(categoria_id=categoria.id, activo=1).first():
            flash('No se puede eliminar la categoría porque tiene productos activos.', 'danger')
        else:
            db.session.delete(categoria)
            db.session.commit()
            flash('Categoría eliminada.', 'warning')
    return redirect(url_for('admin_categorias'))

# ─── Importación Masiva ──────────────────────────────────────
@app.route('/admin/importar', methods=['POST'])
@login_requerido
def admin_importar():
    try:
        if not session.get('admin_autenticado'): 
            return jsonify({"error": "No autorizado"}), 401

        if 'excel_file' not in request.files:
            return jsonify({"error": "No se subió ningún archivo con la clave 'excel_file'."}), 400
        
        file = request.files['excel_file']
        if not file or file.filename == '':
            return jsonify({"error": "Ningún archivo seleccionado."}), 400

        try:
            import openpyxl
        except ImportError:
            raise Exception("La librería 'openpyxl' no está instalada en el servidor. Instálala ejecutando: pip install openpyxl")

        stats = {'actualizados_ok': 0, 'nuevos': 0, 'desactivados': 0}
        wb = openpyxl.load_workbook(file, data_only=True)
        hoja = wb.active
        
        encabezados = {}
        for i, celda in enumerate(hoja[1]):
            if celda.value:
                encabezados[str(celda.value).strip().lower()] = i

        idx_precio = encabezados.get('precio lista 1') or encabezados.get('precio')
        idx_codigo = encabezados.get('código de barras') or encabezados.get('codigo de barras') or encabezados.get('codigo') or encabezados.get('código')
        idx_nombre = encabezados.get('nombre')
        idx_stock = encabezados.get('stock')
        idx_cat = encabezados.get('categoría') or encabezados.get('categoria')
        idx_p2 = encabezados.get('precio_lista_2') or encabezados.get('precio lista 2')
        idx_p3 = encabezados.get('precio_lista_3') or encabezados.get('precio lista 3')

        if idx_precio is None or idx_codigo is None or idx_nombre is None:
            return jsonify({"error": "El Excel debe tener las columnas 'Nombre', 'Precio' y 'Código de barras'."}), 400

        def limpiar_precio(valor_celda):
            if valor_celda is None: return 0.0
            valor_str = str(valor_celda).strip().lower()
            if valor_str in ['nan', 'none', '']: return 0.0
            valor_str = valor_str.replace('$', '').replace(' ', '')
            if ',' in valor_str and '.' not in valor_str: valor_str = valor_str.replace(',', '.')
            elif ',' in valor_str and '.' in valor_str: valor_str = valor_str.replace(',', '')
            try: return float(valor_str)
            except ValueError: return 0.0

        # Eliminamos la carga total de productos en memoria para ahorrar RAM
        ids_procesados = set()
        contador = 0
        BATCH_SIZE = 200

        for fila in hoja.iter_rows(min_row=2):
            if fila[idx_codigo].value is None or fila[idx_nombre].value is None:
                continue
                
            codigo_excel = str(fila[idx_codigo].value).strip()
            if codigo_excel.endswith('.0'): codigo_excel = codigo_excel[:-2]
            
            nombre_excel = str(fila[idx_nombre].value).strip()
            precio_final = limpiar_precio(fila[idx_precio].value)
            
            stock_final = 0
            if idx_stock is not None and fila[idx_stock].value is not None:
                try: stock_final = int(fila[idx_stock].value)
                except: pass

            # UPSERT LOGIC consultando individualmente para no saturar la RAM
            prod = Producto.query.filter_by(codigo_barra=codigo_excel).first()
            if not prod:
                # Intento buscar por nombre si no halló por código
                prod = Producto.query.filter(Producto.nombre.ilike(nombre_excel)).first()

            if prod:
                # Update
                prod.nombre = nombre_excel
                prod.precio_lista_1 = precio_final
                prod.codigo_barra = codigo_excel
                prod.stock = stock_final
                prod.activo = 1
                prod.sincronizado = not es_offline()
                prod.ultima_actualizacion = hora_argentina()
                stats['actualizados_ok'] += 1
            else:
                # Insert
                prod = Producto(
                    nombre=nombre_excel,
                    precio_lista_1=precio_final,
                    codigo_barra=codigo_excel,
                    stock=stock_final,
                    activo=1,
                    sincronizado=not es_offline(),
                    ultima_actualizacion=hora_argentina()
                )
                db.session.add(prod)
                stats['nuevos'] += 1
                
            db.session.flush() # para obtener prod.id si es nuevo
                
            # Resto de campos opcionales
            if idx_cat is not None and fila[idx_cat].value is not None:
                cat_name = str(fila[idx_cat].value).strip()
                if cat_name and cat_name.lower() not in ['nan', 'none']:
                    categoria = Categoria.query.filter(Categoria.nombre.ilike(cat_name)).first()
                    if not categoria:
                        categoria = Categoria(nombre=cat_name)
                        db.session.add(categoria)
                        db.session.flush()
                    prod.categoria_id = categoria.id

            if idx_p2 is not None and fila[idx_p2].value is not None:
                prod.precio_lista_2 = limpiar_precio(fila[idx_p2].value)
            else:
                prod.precio_lista_2 = precio_final
                
            if idx_p3 is not None and fila[idx_p3].value is not None:
                prod.precio_lista_3 = limpiar_precio(fila[idx_p3].value)
            else:
                prod.precio_lista_3 = precio_final

            ids_procesados.add(prod.id)
            contador += 1
            
            # Liberar memoria de SQLAlchemy en lotes
            if contador % BATCH_SIZE == 0:
                db.session.commit()
                db.session.expunge_all()

        # Commit de los elementos restantes
        db.session.commit()
        db.session.expunge_all()

        # SOFT DELETE LOGIC masivo (sin traer objetos a memoria)
        if ids_procesados:
            # Desactivar todo lo que está activo y no está en la lista del Excel
            desactivados = Producto.query.filter(
                Producto.activo == 1,
                ~Producto.id.in_(ids_procesados)
            ).update({
                'activo': 0,
                'ultima_actualizacion': hora_argentina(),
                'sincronizado': not es_offline()
            }, synchronize_session=False)
            stats['desactivados'] = desactivados
            db.session.commit()
            db.session.expunge_all()

        # Garbage Collection forzada para liberar RAM inmediatamente
        import gc
        gc.collect()
        
        global ultima_actualizacion_precios
        ultima_actualizacion_precios = hora_argentina()
        actualizar_version_catalogo()
        
        mensaje = f"✅ Catálogo sincronizado correctamente. Actualizados: {stats['actualizados_ok']}, Nuevos: {stats['nuevos']}, Desactivados: {stats['desactivados']}."
        return jsonify({"mensaje": mensaje}), 200

    except Exception as e:
        db.session.rollback()
        print(f"Error detallado: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/verificar_precios', methods=['GET'])
@login_requerido
def verificar_precios():
    # Si no existe la variable, manda la hora argentina actual como default
    ultima = globals().get('ultima_actualizacion_precios', hora_argentina())
    return jsonify({"ultima_actualizacion": ultima.isoformat()}), 200

@app.route('/api/catalogo/version', methods=['GET'])
def catalogo_version():
    global ultima_actualizacion_catalogo
    return jsonify({"version": ultima_actualizacion_catalogo})

@app.route('/api/catalogo/version/update', methods=['POST'])
def catalogo_version_update():
    actualizar_version_catalogo()
    return jsonify({"ok": True, "version": ultima_actualizacion_catalogo})

# ─── Exportación a Excel ───────────────────────────────────────
@app.route('/admin/exportar')
@login_requerido # Protegido (A-05)
def admin_exportar():
    if not session.get('admin_autenticado'): return redirect('/')
    productos = Producto.query.all()
    data = []
    for p in productos:
        data.append({
            'ID': p.id,
            'Nombre': p.nombre,
            'Precio Lista 1': p.precio_lista_1,
            'Precio Lista 2': p.precio_lista_2,
            'Precio Lista 3': p.precio_lista_3,
            'Código de Barras': p.codigo_barra,
            'Stock': p.stock,
            'Categoria': p.categoria_rel.nombre if p.categoria_rel else 'General',
            'Favorito': 'SI' if p.favorito else 'NO',
            'Venta sin Stock': 'SI' if p.permitir_sin_stock else 'NO',
            'Link Imagen': p.imagen_url or p.imagen,
            'Activo': 'SI' if p.activo == 1 else 'NO'
        })
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row[h] for h in headers])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        download_name='Inventario_Todo_Golosina.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ─── Estadísticas de Ventas ──────────────────────────────────
@app.route('/admin/estadisticas')
@login_requerido # Protegido (A-05)
def admin_estadisticas():
    if not session.get('admin_autenticado'): return redirect('/')
    top_vendidos = Producto.query.filter_by(activo=1).order_by(Producto.ventas_totales.desc()).limit(5).all()
    peor_vendidos = Producto.query.filter_by(activo=1).order_by(Producto.ventas_totales.asc()).limit(5).all()
    
    en_stock = Producto.query.filter(Producto.activo==1, Producto.stock > 0).count()
    sin_stock = Producto.query.filter(Producto.activo==1, Producto.stock <= 0).count()
    total_productos = en_stock + sin_stock
    
    stock_porc = (en_stock / total_productos * 100) if total_productos > 0 else 0
    sin_stock_porc = (sin_stock / total_productos * 100) if total_productos > 0 else 0

    return render_template(
        'admin_estadisticas.html', 
        top_vendidos=top_vendidos,
        peor_vendidos=peor_vendidos,
        en_stock=en_stock,
        sin_stock=sin_stock,
        total_productos=total_productos,
        stock_porc=stock_porc,
        sin_stock_porc=sin_stock_porc
    )

# ─── Configuración inicial de Base de Datos ──────────────────

from sqlalchemy import text

def setup_database():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with app.app_context():
        db.create_all()
        
        # Intentar añadir la columna de código de barras de manera segura por si la tabla ya existe
        try:
            db.session.execute(text('ALTER TABLE "Productos" ADD COLUMN codigo_barra VARCHAR(100);'))
            db.session.commit()
        except Exception:
            db.session.rollback()

        # Agregar columnas de sincronización en Productos de manera segura
        try:
            db.session.execute(text('ALTER TABLE "Productos" ADD COLUMN sincronizado BOOLEAN DEFAULT TRUE;'))
            db.session.commit()
        except Exception:
            db.session.rollback()

        try:
            db.session.execute(text('ALTER TABLE "Productos" ADD COLUMN ultima_actualizacion TIMESTAMP;'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        
        # Columnas nuevas para Clientes (MIGRACIÓN PASO A PASO)
        columnas_clientes = [
            ('cuit', 'VARCHAR(20)'),
            ('condicion_iva', 'VARCHAR(50)'),
            ('descuento_fijo', 'FLOAT DEFAULT 0.0'),
            ('saldo', 'FLOAT DEFAULT 0.0'),
            ('limite_credito', 'FLOAT DEFAULT 0.0')
        ]
        for col, tip in columnas_clientes:
            try:
                db.session.execute(text(f'ALTER TABLE "clientes" ADD COLUMN {col} {tip};'))
                db.session.commit()
            except Exception:
                db.session.rollback()
        
        # Columnas nuevas para Ventas
        columnas_ventas = [
            ('pago_efectivo', 'FLOAT DEFAULT 0.0'),
            ('pago_transferencia', 'FLOAT DEFAULT 0.0'),
            ('pago_debito', 'FLOAT DEFAULT 0.0'),
            ('pago_cc', 'FLOAT DEFAULT 0.0'),
            ('sincronizado', 'BOOLEAN DEFAULT TRUE'),
            ('ultima_actualizacion', 'TIMESTAMP')
        ]
        for col, tip in columnas_ventas:
            try:
                db.session.execute(text(f'ALTER TABLE "ventas" ADD COLUMN {col} {tip};'))
                db.session.commit()
            except Exception:
                db.session.rollback()

        if Categoria.query.count() == 0:
            for cat_name in ['General', 'Gummies', 'Chocolates', 'Chupetines', 'Marshmallows', 'Ácidos', 'Caramelos', 'Gift Boxes']:
                db.session.add(Categoria(nombre=cat_name))
            db.session.commit()
            
            cat_general = Categoria.query.filter_by(nombre='General').first()
            for p in Producto.query.all():
                if not p.categoria_id:
                    p.categoria_id = cat_general.id
            db.session.commit()

        if not Usuario.query.filter_by(username='admin').first():
            admin = Usuario(username='admin', password_hash=generate_password_hash('admin123'))
            db.session.add(admin)
            db.session.commit()

# ─── Inicialización de la Base de Datos ──────────────────────
# Llamamos a esta función aquí para que se ejecute al importar 'app' en Render/Gunicorn
setup_database()

@app.route('/admin/reportes/turnos')
@login_requerido
def reportes_turnos_view():
    if not session.get('admin_autenticado'):
        return redirect('/')
    return render_template('admin_reportes_turnos.html')

@app.route('/api/reportes/turnos', methods=['GET'])
@login_requerido
def api_reportes_turnos():
    from datetime import datetime, time
    import pytz
    
    fecha_str = request.args.get('fecha')
    if not fecha_str:
        tz = pytz.timezone('America/Argentina/Buenos_Aires')
        fecha_str = datetime.now(tz).strftime('%Y-%m-%d')
        
    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({"ok": False, "mensaje": "Fecha inválida"}), 400
        
    ventas = Venta.query.filter(
        Venta.fecha >= datetime.combine(fecha_obj, time.min),
        Venta.fecha <= datetime.combine(fecha_obj, time.max)
    ).all()
    
    # Agrupar por turno
    turnos_stats = {
        'Mañana': {'total_recaudado': 0.0, 'cantidad_ventas': 0, 'ventas': []},
        'Tarde': {'total_recaudado': 0.0, 'cantidad_ventas': 0, 'ventas': []}
    }
    
    for v in ventas:
        hora = v.fecha.hour
        turno = 'Mañana' if hora < 13 else 'Tarde'
        
        turnos_stats[turno]['total_recaudado'] += v.total
        turnos_stats[turno]['cantidad_ventas'] += 1
        turnos_stats[turno]['ventas'].append(v)
        
    # Calcular productos más vendidos por turno
    import json
    for turno_key in turnos_stats.keys():
        productos_count = {}
        for v in turnos_stats[turno_key]['ventas']:
            try:
                detalle = json.loads(v.detalle_json)
                for item in detalle:
                    nombre = item.get('nombre', 'Desconocido')
                    qty = int(item.get('qty', 1))
                    productos_count[nombre] = productos_count.get(nombre, 0) + qty
            except Exception:
                pass
                
        # Ordenar y tomar los 5 principales
        top_productos = sorted(productos_count.items(), key=lambda x: x[1], reverse=True)[:5]
        turnos_stats[turno_key]['productos_top'] = [{"nombre": k, "cantidad": v} for k, v in top_productos]
        
        # Eliminamos array ventas para no saturar json
        del turnos_stats[turno_key]['ventas']
        
    return jsonify({
        "ok": True,
        "fecha": fecha_str,
        "turnos": turnos_stats
    })

@app.route('/reportes')
@login_requerido # Protegido (A-05)
def reportes_view():
    if not session.get('admin_autenticado'):
        return redirect('/')
    return render_template('reportes.html')

@app.route('/api/reportes', methods=['GET'])
@login_requerido # Protegido (A-05)
def api_reportes():
    from datetime import datetime, time
    import json
    
    fecha_desde_str = request.args.get('desde')
    fecha_hasta_str = request.args.get('hasta')
    
    query = Venta.query
    
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
            query = query.filter(Venta.fecha >= datetime.combine(fecha_desde, time.min))
        except ValueError:
            pass
    
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d')
            query = query.filter(Venta.fecha <= datetime.combine(fecha_hasta, time.max))
        except ValueError:
            pass
        
    ventas = query.all()
    
    resumen = {
        "Efectivo": 0.0,
        "Transferencia": 0.0,
        "Débito": 0.0,
        "Total": 0.0
    }
    
    clientes_reporte = {}
    
    for v in ventas:
        resumen["Efectivo"] += (v.pago_efectivo or 0.0)
        resumen["Transferencia"] += (v.pago_transferencia or 0.0)
        resumen["Débito"] += (v.pago_debito or 0.0)
        resumen["Total"] += v.total
        
        # Reporte por Clientes
        c_id = v.cliente_id or 0
        c_nombre = v.cliente.nombre if v.cliente else 'Consumidor Final'
        
        if c_id not in clientes_reporte:
            clientes_reporte[c_id] = {"nombre": c_nombre, "total": 0.0}
        
        clientes_reporte[c_id]["total"] += v.total
        
    clientes_lista = sorted(
        [{"id": cid, "nombre": data["nombre"], "total": data["total"]} for cid, data in clientes_reporte.items()],
        key=lambda x: x["total"],
        reverse=True
    )
    
    return jsonify({
        "ok": True,
        "resumen": resumen,
        "clientes": clientes_lista
    })

@app.route('/manifest.json')
@login_requerido
def send_manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
@login_requerido
def send_sw():
    response = make_response(send_from_directory('static', 'sw.js'))
    response.headers['Content-Type'] = 'application/javascript'
    return response

if __name__ == '__main__':
    app.run(debug=True, port=5000)